import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
wb.remove(wb.active)

def setup_sheet(ws, headers, title):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(col_num)].width = 18

companies = ['Tân Hồng Hà', 'Việt', 'Xem Sơn', 'VPS M', 'ITSS', 'Văn phòng VPS']

# 1. Doanh thu
ws_rev = wb.create_sheet('Doanh thu')
rev_headers = ['Công ty', 'Tháng', 'Doanh số Kế hoạch (VNĐ)', 'Doanh số Thực tế (VNĐ)', 'Tỷ lệ gộp', 'Chi phí (VNĐ)', 'Lợi nhuận TT (VNĐ)']
setup_sheet(ws_rev, rev_headers, 'BÁO CÁO DOANH THU & LỢI NHUẬN')
row = 3
for comp in companies:
    ws_rev.cell(row=row, column=1, value=comp)
    row += 1

# 2. Công nợ
ws_debt = wb.create_sheet('Công nợ')
debt_headers = ['Công ty', 'Tháng', 'Phải thu trong hạn (VNĐ)', 'Nợ quá hạn (VNĐ)', 'Nợ khó đòi (VNĐ)']
setup_sheet(ws_debt, debt_headers, 'BÁO CÁO CÔNG NỢ')
row = 3
for comp in companies:
    ws_debt.cell(row=row, column=1, value=comp)
    row += 1

# 3. Khách hàng
ws_cust = wb.create_sheet('Khách hàng')
# We need a complex header for Customers
ws_cust.merge_cells('A1:L1')
ws_cust['A1'] = 'BÁO CÁO CƠ CẤU KHÁCH HÀNG'
ws_cust['A1'].font = Font(size=14, bold=True)
ws_cust['A1'].alignment = Alignment(horizontal='center', vertical='center')

headers1 = ['Công ty', 'Mảng kinh doanh', 'Đầu kỳ', '', 'Tăng trong kỳ', '', 'Giảm trong kỳ', '', 'Cuối kỳ', '']
headers2 = ['', '', 'Số Máy', 'Số KH', 'Số Máy', 'Số KH', 'Số Máy', 'Số KH', 'Số Máy', 'Số KH']
for col, val in enumerate(headers1, 1):
    cell = ws_cust.cell(row=2, column=col, value=val)
    cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    cell.font = Font(color="FFFFFF", bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
for col, val in enumerate(headers2, 1):
    cell = ws_cust.cell(row=3, column=col, value=val)
    cell.fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
    cell.font = Font(color="FFFFFF", bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_cust.column_dimensions[get_column_letter(col)].width = 15

# Merges for headers1
ws_cust.merge_cells('A2:A3')
ws_cust.merge_cells('B2:B3')
ws_cust.merge_cells('C2:D2')
ws_cust.merge_cells('E2:F2')
ws_cust.merge_cells('G2:H2')
ws_cust.merge_cells('I2:J2')

categories = ['Thuê máy', 'MC', 'Dịch vụ - Photo', 'Dịch vụ - Máy in', 'Dịch vụ khác', 'Phân phối (Đại lý)']
row = 4
for comp in companies:
    start_r = row
    for cat in categories:
        ws_cust.cell(row=row, column=1, value=comp)
        ws_cust.cell(row=row, column=2, value=cat)
        if cat == 'Phân phối (Đại lý)':
            ws_cust.cell(row=row, column=3, value='-') # No Machine count
            ws_cust.cell(row=row, column=5, value='-')
            ws_cust.cell(row=row, column=7, value='-')
            ws_cust.cell(row=row, column=9, value='-')
        row += 1
    ws_cust.merge_cells(f'A{start_r}:A{row-1}')
    ws_cust[f'A{start_r}'].alignment = Alignment(vertical='center')

# 4. Tồn kho
ws_inv = wb.create_sheet('Tồn kho')
inv_headers = ['Công ty', 'Danh mục', 'Tên Vật tư/Thiết bị', 'Đơn vị tính', 'Số lượng', 'Tổng Giá trị (VNĐ)']
setup_sheet(ws_inv, inv_headers, 'BÁO CÁO TỒN KHO')
ws_inv.column_dimensions['C'].width = 30
row = 3
for comp in companies:
    ws_inv.cell(row=row, column=1, value=comp)
    row += 1

# 5. Nhân sự
ws_hr = wb.create_sheet('Nhân sự')
hr_headers = ['Công ty', 'Phòng ban', 'Tổng NV Đầu kỳ', 'Tuyển mới', 'Nghỉ việc', 'NV Thử việc', 'Tổng NV Cuối kỳ']
setup_sheet(ws_hr, hr_headers, 'BÁO CÁO NHÂN SỰ')
depts = ['Kinh doanh', 'Kỹ thuật', 'Kế toán', 'Hành chính', 'Kho/Giao vận']
row = 3
for comp in companies:
    start_r = row
    for dept in depts:
        ws_hr.cell(row=row, column=1, value=comp)
        ws_hr.cell(row=row, column=2, value=dept)
        row += 1
    ws_hr.merge_cells(f'A{start_r}:A{row-1}')
    ws_hr[f'A{start_r}'].alignment = Alignment(vertical='center')

artifact_path = r'C:\Users\Hp\.gemini\antigravity\brain\c2c07afe-a201-42e1-b52e-07538b848a35\Template_Nhap_Lieu_VPS_V2.xlsx'
wb.save(artifact_path)
print(f"Saved {artifact_path}")
